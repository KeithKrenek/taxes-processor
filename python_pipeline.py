"""
Financial Transaction Processing Pipeline
========================================

A comprehensive system for processing financial transactions with ML-powered
categorization, duplicate detection, and advanced analytics.

Author: Portfolio Demonstration
Target Roles: AI Researcher, Applied Physicist, Technical Lead, Engineering Manager
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from typing import Dict, List, Tuple, Optional, Any
import yaml
from dataclasses import dataclass, field
from datetime import datetime, timedelta
import re
from collections import defaultdict

# ML and Data Science Libraries
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.naive_bayes import MultinomialNB
from sklearn.svm import SVC
from sklearn.cluster import KMeans
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.decomposition import PCA
from sklearn.pipeline import Pipeline

# Fuzzy matching and text processing
from fuzzywuzzy import fuzz, process
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import PorterStemmer

# Visualization libraries
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.figure_factory as ff

# Excel processing
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

# Statistical analysis
from scipy import stats
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.spatial.distance import pdist


@dataclass
class TransactionRecord:
    """Structured representation of a financial transaction."""
    date: datetime
    amount: float
    description: str
    account: str
    category: Optional[str] = None
    confidence: float = 0.0
    is_duplicate: bool = False
    duplicate_group: Optional[int] = None
    source_sheet: str = ""
    raw_data: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ProcessingConfig:
    """Configuration parameters for the processing pipeline."""
    
    # Categorization settings
    ml_confidence_threshold: float = 0.8
    enable_fuzzy_matching: bool = True
    fuzzy_threshold: int = 85
    
    # Duplicate detection settings
    temporal_window_days: int = 5
    amount_tolerance: float = 0.01
    similarity_threshold: float = 0.85
    
    # ML model settings
    enable_ensemble: bool = True
    cross_validation_folds: int = 5
    test_size: float = 0.2
    random_state: int = 42
    
    # Visualization settings
    chart_theme: str = "plotly_white"
    export_formats: List[str] = field(default_factory=lambda: ["html", "png"])
    
    @classmethod
    def from_yaml(cls, config_path: str) -> 'ProcessingConfig':
        """Load configuration from YAML file."""
        with open(config_path, 'r') as f:
            config_data = yaml.safe_load(f)
        return cls(**config_data)


class DataLoader:
    """Advanced Excel data loading with validation and preprocessing."""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
    def load_workbook(self, file_path: str) -> List[TransactionRecord]:
        """Load and validate Excel workbook containing financial transactions."""
        self.logger.info(f"Loading workbook: {file_path}")
        
        try:
            # Read Excel file with multiple sheets
            excel_file = pd.ExcelFile(file_path)
            transactions = []
            
            # Process each sheet (excluding category reference sheets)
            transaction_sheets = [sheet for sheet in excel_file.sheet_names 
                                if not sheet.lower().endswith('categories')]
            
            for sheet_name in transaction_sheets:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_transactions = self._process_sheet(df, sheet_name)
                transactions.extend(sheet_transactions)
                
            self.logger.info(f"Loaded {len(transactions)} transactions from {len(transaction_sheets)} sheets")
            return transactions
            
        except Exception as e:
            self.logger.error(f"Error loading workbook: {e}")
            raise
    
    def _process_sheet(self, df: pd.DataFrame, sheet_name: str) -> List[TransactionRecord]:
        """Process individual sheet data into TransactionRecord objects."""
        transactions = []
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()
        column_mapping = self._map_columns(df.columns)
        
        for idx, row in df.iterrows():
            try:
                # Extract core transaction data
                transaction = self._create_transaction_record(row, column_mapping, sheet_name)
                if transaction:
                    transactions.append(transaction)
            except Exception as e:
                self.logger.warning(f"Error processing row {idx} in sheet {sheet_name}: {e}")
                
        return transactions
    
    def _map_columns(self, columns: List[str]) -> Dict[str, str]:
        """Map sheet columns to standard transaction fields."""
        mapping = {}
        
        # Date field mapping
        date_patterns = ['date', 'posting date', 'transaction date', 'trans date']
        mapping['date'] = self._find_column_match(columns, date_patterns)
        
        # Amount field mapping
        amount_patterns = ['amount', 'transaction amount', 'debit', 'credit']
        mapping['amount'] = self._find_column_match(columns, amount_patterns)
        
        # Description field mapping
        desc_patterns = ['description', 'memo', 'details', 'transaction details']
        mapping['description'] = self._find_column_match(columns, desc_patterns)
        
        return mapping
    
    def _find_column_match(self, columns: List[str], patterns: List[str]) -> Optional[str]:
        """Find best matching column name using fuzzy matching."""
        best_match = None
        best_score = 0
        
        for pattern in patterns:
            for col in columns:
                score = fuzz.ratio(pattern, col)
                if score > best_score and score > 70:
                    best_score = score
                    best_match = col
                    
        return best_match
    
    def _create_transaction_record(self, row: pd.Series, mapping: Dict[str, str], 
                                 sheet_name: str) -> Optional[TransactionRecord]:
        """Create TransactionRecord from pandas row."""
        try:
            # Extract date
            date_col = mapping.get('date')
            if date_col and pd.notna(row[date_col]):
                date = pd.to_datetime(row[date_col])
            else:
                return None
            
            # Extract amount
            amount_col = mapping.get('amount')
            if amount_col and pd.notna(row[amount_col]):
                amount = float(row[amount_col])
            else:
                return None
            
            # Extract description
            desc_col = mapping.get('description')
            description = str(row[desc_col]) if desc_col and pd.notna(row[desc_col]) else ""
            
            # Clean description
            description = self._clean_description(description)
            
            return TransactionRecord(
                date=date,
                amount=amount,
                description=description,
                account=sheet_name,
                source_sheet=sheet_name,
                raw_data=row.to_dict()
            )
            
        except Exception as e:
            self.logger.warning(f"Error creating transaction record: {e}")
            return None
    
    def _clean_description(self, description: str) -> str:
        """Clean and normalize transaction descriptions."""
        if not description:
            return ""
        
        # Remove extra whitespace
        description = re.sub(r'\s+', ' ', description).strip()
        
        # Remove non-printable characters
        description = re.sub(r'[^\x20-\x7E]', '', description)
        
        # Normalize common patterns
        description = re.sub(r'\b(purchase|payment|transfer)\b', '', description, flags=re.IGNORECASE)
        
        return description


class MLCategorizer:
    """Machine learning-powered transaction categorization system."""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # Initialize ML components
        self.vectorizer = TfidfVectorizer(
            max_features=5000,
            stop_words='english',
            ngram_range=(1, 2),
            min_df=2
        )
        
        self.label_encoder = LabelEncoder()
        self.scaler = StandardScaler()
        
        # Ensemble classifier
        self.classifier = VotingClassifier([
            ('rf', RandomForestClassifier(n_estimators=100, random_state=config.random_state)),
            ('nb', MultinomialNB()),
            ('svm', SVC(probability=True, random_state=config.random_state))
        ], voting='soft')
        
        # Category mappings
        self.category_keywords = self._load_category_keywords()
        self.trained = False
        
    def _load_category_keywords(self) -> Dict[str, List[str]]:
        """Load predefined category keywords and patterns."""
        return {
            "Food & Dining": [
                "restaurant", "cafe", "coffee", "pizza", "burger", "taco", "doordash", 
                "uber eats", "grubhub", "dining", "food", "kitchen", "deli", "bakery"
            ],
            "Transportation": [
                "uber", "lyft", "taxi", "gas", "fuel", "parking", "metro", "bus", 
                "train", "airline", "flight", "rental car", "auto", "vehicle"
            ],
            "Shopping": [
                "amazon", "target", "walmart", "costco", "store", "retail", "shopping",
                "clothing", "apparel", "electronics", "purchase"
            ],
            "Utilities": [
                "electric", "water", "gas", "internet", "phone", "cable", "utility",
                "bill", "service", "provider", "connection"
            ],
            "Healthcare": [
                "medical", "doctor", "pharmacy", "hospital", "clinic", "health",
                "dental", "vision", "prescription", "copay"
            ],
            "Entertainment": [
                "movie", "theater", "concert", "streaming", "netflix", "spotify",
                "entertainment", "game", "recreation", "hobby"
            ],
            "Financial Services": [
                "bank", "atm", "fee", "interest", "loan", "credit", "investment",
                "financial", "advisor", "brokerage"
            ],
            "Home & Garden": [
                "home depot", "lowes", "garden", "hardware", "improvement", "repair",
                "maintenance", "contractor", "furniture"
            ],
            "Professional Services": [
                "legal", "accounting", "consulting", "professional", "service",
                "attorney", "cpa", "advisor"
            ],
            "Education": [
                "school", "university", "college", "education", "tuition", "books",
                "course", "training", "certification"
            ]
        }
    
    def extract_features(self, transactions: List[TransactionRecord]) -> np.ndarray:
        """Extract features for machine learning classification."""
        features = []
        
        for transaction in transactions:
            # Text features
            text_features = self._extract_text_features(transaction.description)
            
            # Amount features
            amount_features = self._extract_amount_features(transaction.amount)
            
            # Temporal features
            temporal_features = self._extract_temporal_features(transaction.date)
            
            # Combine all features
            combined_features = {**text_features, **amount_features, **temporal_features}
            features.append(combined_features)
        
        return pd.DataFrame(features)
    
    def _extract_text_features(self, description: str) -> Dict[str, float]:
        """Extract text-based features from transaction description."""
        description = description.lower()
        
        features = {
            'length': len(description),
            'word_count': len(description.split()),
            'has_digits': float(any(c.isdigit() for c in description)),
            'has_uppercase': float(any(c.isupper() for c in description)),
        }
        
        # Keyword presence features
        for category, keywords in self.category_keywords.items():
            features[f'keyword_{category.lower()}'] = float(
                any(keyword in description for keyword in keywords)
            )
        
        return features
    
    def _extract_amount_features(self, amount: float) -> Dict[str, float]:
        """Extract amount-based features."""
        return {
            'amount': amount,
            'amount_log': np.log1p(abs(amount)),
            'is_round': float(amount == round(amount)),
            'is_negative': float(amount < 0),
            'amount_range': self._categorize_amount(amount)
        }
    
    def _extract_temporal_features(self, date: datetime) -> Dict[str, float]:
        """Extract temporal features from transaction date."""
        return {
            'day_of_week': float(date.weekday()),
            'day_of_month': float(date.day),
            'month': float(date.month),
            'quarter': float((date.month - 1) // 3 + 1),
            'is_weekend': float(date.weekday() >= 5),
            'is_month_end': float(date.day > 25)
        }
    
    def _categorize_amount(self, amount: float) -> float:
        """Categorize amount into ranges."""
        abs_amount = abs(amount)
        if abs_amount < 10:
            return 1
        elif abs_amount < 50:
            return 2
        elif abs_amount < 200:
            return 3
        elif abs_amount < 1000:
            return 4
        else:
            return 5
    
    def train(self, transactions: List[TransactionRecord], labels: List[str]):
        """Train the ML categorization model."""
        self.logger.info("Training ML categorization model...")
        
        # Extract features
        feature_df = self.extract_features(transactions)
        
        # Prepare text data for TF-IDF
        descriptions = [t.description for t in transactions]
        text_features = self.vectorizer.fit_transform(descriptions).toarray()
        
        # Combine numerical and text features
        numerical_features = self.scaler.fit_transform(feature_df.values)
        combined_features = np.hstack([numerical_features, text_features])
        
        # Encode labels
        encoded_labels = self.label_encoder.fit_transform(labels)
        
        # Train classifier
        self.classifier.fit(combined_features, encoded_labels)
        
        # Evaluate model
        scores = cross_val_score(self.classifier, combined_features, encoded_labels, 
                               cv=self.config.cross_validation_folds)
        self.logger.info(f"Cross-validation accuracy: {scores.mean():.3f} (+/- {scores.std() * 2:.3f})")
        
        self.trained = True
    
    def predict(self, transactions: List[TransactionRecord]) -> List[Tuple[str, float]]:
        """Predict categories for transactions."""
        if not self.trained:
            raise ValueError("Model must be trained before making predictions")
        
        # Extract features
        feature_df = self.extract_features(transactions)
        
        # Prepare text data
        descriptions = [t.description for t in transactions]
        text_features = self.vectorizer.transform(descriptions).toarray()
        
        # Combine features
        numerical_features = self.scaler.transform(feature_df.values)
        combined_features = np.hstack([numerical_features, text_features])
        
        # Make predictions
        predictions = self.classifier.predict(combined_features)
        probabilities = self.classifier.predict_proba(combined_features)
        
        results = []
        for i, pred in enumerate(predictions):
            category = self.label_encoder.inverse_transform([pred])[0]
            confidence = np.max(probabilities[i])
            results.append((category, confidence))
        
        return results
    
    def categorize_transactions(self, transactions: List[TransactionRecord]) -> List[TransactionRecord]:
        """Categorize transactions using rule-based and ML approaches."""
        categorized = []
        
        for transaction in transactions:
            # Try rule-based categorization first
            category, confidence = self._rule_based_categorization(transaction)
            
            # If rule-based confidence is low and ML is available, use ML
            if confidence < self.config.ml_confidence_threshold and self.trained:
                ml_category, ml_confidence = self.predict([transaction])[0]
                if ml_confidence > confidence:
                    category, confidence = ml_category, ml_confidence
            
            # Update transaction
            transaction.category = category
            transaction.confidence = confidence
            categorized.append(transaction)
        
        return categorized
    
    def _rule_based_categorization(self, transaction: TransactionRecord) -> Tuple[str, float]:
        """Apply rule-based categorization logic."""
        description = transaction.description.lower()
        amount = transaction.amount
        
        # High-confidence keyword matches
        for category, keywords in self.category_keywords.items():
            for keyword in keywords:
                if keyword in description:
                    # Use fuzzy matching for additional confidence
                    similarity = max(fuzz.partial_ratio(keyword, description) for keyword in keywords)
                    confidence = min(0.95, similarity / 100.0)
                    return category, confidence
        
        # Amount-based rules
        if abs(amount) > 1000:
            return "Large Purchase", 0.6
        elif amount < 0 and abs(amount) < 10:
            return "Small Expense", 0.7
        
        return "Uncategorized", 0.1


class DuplicateDetector:
    """Advanced duplicate detection using multiple similarity metrics."""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
    
    def detect_duplicates(self, transactions: List[TransactionRecord]) -> List[TransactionRecord]:
        """Detect and mark duplicate transactions."""
        self.logger.info("Detecting duplicate transactions...")
        
        # Sort transactions by date for efficient processing
        sorted_transactions = sorted(transactions, key=lambda x: x.date)
        duplicate_groups = []
        processed_indices = set()
        
        for i, transaction in enumerate(sorted_transactions):
            if i in processed_indices:
                continue
            
            # Find potential duplicates within temporal window
            duplicates = self._find_duplicate_candidates(transaction, sorted_transactions[i+1:], i+1)
            
            if duplicates:
                group_id = len(duplicate_groups)
                duplicate_groups.append([i] + duplicates)
                
                # Mark all transactions in group as duplicates
                for idx in [i] + duplicates:
                    sorted_transactions[idx].is_duplicate = True
                    sorted_transactions[idx].duplicate_group = group_id
                    processed_indices.add(idx)
        
        self.logger.info(f"Found {len(duplicate_groups)} duplicate groups affecting {len(processed_indices)} transactions")
        return sorted_transactions
    
    def _find_duplicate_candidates(self, target: TransactionRecord, 
                                 candidates: List[TransactionRecord], 
                                 start_idx: int) -> List[int]:
        """Find duplicate candidates for a given transaction."""
        duplicates = []
        
        for i, candidate in enumerate(candidates):
            actual_idx = start_idx + i
            
            # Check temporal proximity
            if abs((target.date - candidate.date).days) > self.config.temporal_window_days:
                continue
            
            # Check amount similarity
            if abs(target.amount - candidate.amount) > self.config.amount_tolerance:
                continue
            
            # Check description similarity
            description_similarity = self._calculate_description_similarity(
                target.description, candidate.description
            )
            
            if description_similarity >= self.config.similarity_threshold:
                duplicates.append(actual_idx)
        
        return duplicates
    
    def _calculate_description_similarity(self, desc1: str, desc2: str) -> float:
        """Calculate similarity between transaction descriptions."""
        # Normalize descriptions
        desc1_norm = self._normalize_description(desc1)
        desc2_norm = self._normalize_description(desc2)
        
        # Calculate multiple similarity metrics
        token_ratio = fuzz.token_sort_ratio(desc1_norm, desc2_norm)
        partial_ratio = fuzz.partial_ratio(desc1_norm, desc2_norm)
        ratio = fuzz.ratio(desc1_norm, desc2_norm)
        
        # Weighted combination
        combined_score = (token_ratio * 0.4 + partial_ratio * 0.4 + ratio * 0.2) / 100.0
        
        return combined_score
    
    def _normalize_description(self, description: str) -> str:
        """Normalize description for comparison."""
        # Convert to lowercase
        normalized = description.lower()
        
        # Remove common transaction artifacts
        patterns_to_remove = [
            r'\b\d{4,}\b',  # Long numbers (card numbers, reference numbers)
            r'\b\d{1,2}\/\d{1,2}\/\d{2,4}\b',  # Dates
            r'\b\d{1,2}:\d{2}(:\d{2})?\b',  # Times
            r'#\d+',  # Reference numbers
            r'\bpos\b', r'\batm\b',  # Transaction types
        ]
        
        for pattern in patterns_to_remove:
            normalized = re.sub(pattern, '', normalized)
        
        # Remove extra whitespace
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        
        return normalized


class FinancialVisualizer:
    """Professional visualization suite for financial data analysis."""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # Set theme
        self.theme = config.chart_theme
    
    def create_comprehensive_dashboard(self, transactions: List[TransactionRecord], 
                                    output_path: str = "financial_dashboard.html"):
        """Create comprehensive financial analysis dashboard."""
        self.logger.info("Creating comprehensive financial dashboard...")
        
        # Convert to DataFrame for easier manipulation
        df = self._transactions_to_dataframe(transactions)
        
        # Create subplots
        fig = make_subplots(
            rows=3, cols=2,
            subplot_titles=[
                'Monthly Spending Trends',
                'Category Distribution', 
                'Transaction Amount Distribution',
                'Duplicate Detection Analysis',
                'Daily Transaction Patterns',
                'Category Confidence Scores'
            ],
            specs=[
                [{"secondary_y": True}, {"type": "pie"}],
                [{"type": "histogram"}, {"type": "bar"}],
                [{"type": "heatmap"}, {"type": "box"}]
            ]
        )
        
        # Monthly trends
        monthly_data = self._prepare_monthly_data(df)
        fig.add_trace(
            go.Scatter(
                x=monthly_data.index,
                y=monthly_data['amount'],
                mode='lines+markers',
                name='Monthly Spending',
                line=dict(color='#1f77b4', width=3)
            ),
            row=1, col=1
        )
        
        # Category distribution
        category_data = df.groupby('category')['amount'].sum().abs()
        fig.add_trace(
            go.Pie(
                labels=category_data.index,
                values=category_data.values,
                name="Category Distribution"
            ),
            row=1, col=2
        )
        
        # Amount distribution
        fig.add_trace(
            go.Histogram(
                x=df['amount'].abs(),
                nbinsx=50,
                name='Amount Distribution',
                marker_color='#ff7f0e'
            ),
            row=2, col=1
        )
        
        # Duplicate analysis
        duplicate_stats = self._calculate_duplicate_stats(df)
        fig.add_trace(
            go.Bar(
                x=list(duplicate_stats.keys()),
                y=list(duplicate_stats.values()),
                name='Duplicate Analysis',
                marker_color='#d62728'
            ),
            row=2, col=2
        )
        
        # Daily patterns heatmap
        daily_patterns = self._prepare_daily_patterns(df)
        fig.add_trace(
            go.Heatmap(
                z=daily_patterns.values,
                x=daily_patterns.columns,
                y=daily_patterns.index,
                colorscale='Viridis',
                name='Daily Patterns'
            ),
            row=3, col=1
        )
        
        # Confidence scores
        confidence_data = df.groupby('category')['confidence'].mean()
        fig.add_trace(
            go.Box(
                y=df['confidence'],
                x=df['category'],
                name='Confidence Scores',
                marker_color='#2ca02c'
            ),
            row=3, col=2
        )
        
        # Update layout
        fig.update_layout(
            height=1200,
            title_text="Comprehensive Financial Analysis Dashboard",
            template=self.theme,
            showlegend=False
        )
        
        # Save dashboard
        fig.write_html(output_path)
        self.logger.info(f"Dashboard saved to {output_path}")
        
        return fig
    
    def _transactions_to_dataframe(self, transactions: List[TransactionRecord]) -> pd.DataFrame:
        """Convert transaction records to pandas DataFrame."""
        data = []
        for t in transactions:
            data.append({
                'date': t.date,
                'amount': t.amount,
                'description': t.description,
                'category': t.category or 'Uncategorized',
                'confidence': t.confidence,
                'is_duplicate': t.is_duplicate,
                'account': t.account
            })
        return pd.DataFrame(data)
    
    def _prepare_monthly_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare monthly aggregated data."""
        df['month'] = df['date'].dt.to_period('M')
        return df.groupby('month').agg({
            'amount': 'sum',
            'date': 'count'
        }).rename(columns={'date': 'transaction_count'})
    
    def _calculate_duplicate_stats(self, df: pd.DataFrame) -> Dict[str, int]:
        """Calculate duplicate detection statistics."""
        return {
            'Total Transactions': len(df),
            'Unique Transactions': len(df[~df['is_duplicate']]),
            'Duplicate Transactions': len(df[df['is_duplicate']]),
            'Duplicate Groups': df[df['is_duplicate']]['duplicate_group'].nunique()
        }
    
    def _prepare_daily_patterns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare daily transaction patterns for heatmap."""
        df['day_of_week'] = df['date'].dt.day_name()
        df['hour'] = df['date'].dt.hour
        
        # Create pivot table for heatmap
        pivot = df.groupby(['day_of_week', 'hour']).size().unstack(fill_value=0)
        
        # Reorder days
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        pivot = pivot.reindex(day_order)
        
        return pivot


class FinancialPipeline:
    """Main orchestration pipeline for financial transaction processing."""
    
    def __init__(self, config_path: Optional[str] = None):
        # Setup logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
        
        # Load configuration
        if config_path:
            self.config = ProcessingConfig.from_yaml(config_path)
        else:
            self.config = ProcessingConfig()
        
        # Initialize components
        self.data_loader = DataLoader(self.config)
        self.categorizer = MLCategorizer(self.config)
        self.duplicate_detector = DuplicateDetector(self.config)
        self.visualizer = FinancialVisualizer(self.config)
        
        self.logger.info("Financial pipeline initialized successfully")
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """Process a financial data file through the complete pipeline."""
        self.logger.info(f"Starting processing pipeline for: {file_path}")
        
        start_time = datetime.now()
        
        # Load data
        transactions = self.data_loader.load_workbook(file_path)
        
        # Categorize transactions
        categorized_transactions = self.categorizer.categorize_transactions(transactions)
        
        # Detect duplicates
        final_transactions = self.duplicate_detector.detect_duplicates(categorized_transactions)
        
        # Generate summary statistics
        summary = self._generate_summary(final_transactions)
        
        processing_time = datetime.now() - start_time
        self.logger.info(f"Processing completed in {processing_time.total_seconds():.2f} seconds")
        
        return {
            'transactions': final_transactions,
            'summary': summary,
            'processing_time': processing_time,
            'config': self.config
        }
    
    def generate_reports(self, results: Dict[str, Any], output_dir: str = "output/"):
        """Generate comprehensive reports and visualizations."""
        self.logger.info("Generating reports and visualizations...")
        
        # Create output directory
        Path(output_dir).mkdir(exist_ok=True)
        
        transactions = results['transactions']
        
        # Generate dashboard
        dashboard_path = Path(output_dir) / "financial_dashboard.html"
        self.visualizer.create_comprehensive_dashboard(transactions, str(dashboard_path))
        
        # Export processed data
        self._export_processed_data(transactions, output_dir)
        
        # Generate summary report
        self._generate_summary_report(results, output_dir)
        
        self.logger.info(f"Reports generated in {output_dir}")
    
    def _generate_summary(self, transactions: List[TransactionRecord]) -> Dict[str, Any]:
        """Generate comprehensive summary statistics."""
        df = pd.DataFrame([{
            'date': t.date,
            'amount': t.amount,
            'category': t.category,
            'confidence': t.confidence,
            'is_duplicate': t.is_duplicate
        } for t in transactions])
        
        return {
            'total_transactions': len(transactions),
            'unique_transactions': len(df[~df['is_duplicate']]),
            'duplicate_count': len(df[df['is_duplicate']]),
            'total_amount': df['amount'].sum(),
            'date_range': {
                'start': df['date'].min(),
                'end': df['date'].max()
            },
            'categories': {
                'count': df['category'].nunique(),
                'distribution': df.groupby('category')['amount'].sum().to_dict()
            },
            'confidence_stats': {
                'mean': df['confidence'].mean(),
                'median': df['confidence'].median(),
                'std': df['confidence'].std()
            }
        }
    
    def _export_processed_data(self, transactions: List[TransactionRecord], output_dir: str):
        """Export processed transaction data to Excel."""
        df = pd.DataFrame([{
            'Date': t.date,
            'Amount': t.amount,
            'Description': t.description,
            'Category': t.category,
            'Confidence': t.confidence,
            'Is_Duplicate': t.is_duplicate,
            'Duplicate_Group': t.duplicate_group,
            'Account': t.account,
            'Source_Sheet': t.source_sheet
        } for t in transactions])
        
        output_path = Path(output_dir) / "processed_transactions.xlsx"
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Processed_Transactions', index=False)
            
            # Summary by category
            category_summary = df.groupby('Category').agg({
                'Amount': ['sum', 'count', 'mean'],
                'Confidence': 'mean'
            }).round(2)
            category_summary.to_excel(writer, sheet_name='Category_Summary')
            
            # Duplicate analysis
            if df['Is_Duplicate'].any():
                duplicates = df[df['Is_Duplicate']]
                duplicates.to_excel(writer, sheet_name='Duplicates', index=False)
        
        self.logger.info(f"Processed data exported to {output_path}")
    
    def _generate_summary_report(self, results: Dict[str, Any], output_dir: str):
        """Generate text-based summary report."""
        summary = results['summary']
        
        report = f"""
Financial Transaction Processing Report
======================================

Processing Summary:
- Total Transactions Processed: {summary['total_transactions']:,}
- Unique Transactions: {summary['unique_transactions']:,}
- Duplicate Transactions: {summary['duplicate_count']:,}
- Processing Time: {results['processing_time'].total_seconds():.2f} seconds

Financial Overview:
- Total Amount: ${summary['total_amount']:,.2f}
- Date Range: {summary['date_range']['start'].strftime('%Y-%m-%d')} to {summary['date_range']['end'].strftime('%Y-%m-%d')}
- Categories Identified: {summary['categories']['count']}

Categorization Performance:
- Average Confidence: {summary['confidence_stats']['mean']:.2f}
- Median Confidence: {summary['confidence_stats']['median']:.2f}
- Confidence Standard Deviation: {summary['confidence_stats']['std']:.2f}

Top Categories by Amount:
"""
        
        # Add top categories
        sorted_categories = sorted(summary['categories']['distribution'].items(), 
                                 key=lambda x: abs(x[1]), reverse=True)
        
        for category, amount in sorted_categories[:10]:
            report += f"- {category}: ${amount:,.2f}\n"
        
        # Save report
        report_path = Path(output_dir) / "processing_report.txt"
        with open(report_path, 'w') as f:
            f.write(report)
        
        self.logger.info(f"Summary report saved to {report_path}")


# Example usage and demonstration
if __name__ == "__main__":
    # Initialize pipeline
    pipeline = FinancialPipeline()
    
    # Process sample file
    try:
        results = pipeline.process_file("data/sample/financial_data.xlsx")
        pipeline.generate_reports(results)
        
        print("Processing completed successfully!")
        print(f"Summary: {results['summary']}")
        
    except FileNotFoundError:
        print("Sample data file not found. Please provide a valid Excel file path.")
    except Exception as e:
        print(f"Error during processing: {e}")
